from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import os

def scrivi_statistiche_excel(file_path, soglia_giocatore, dati):
    """
    Scrive le statistiche in un file Excel.
    La prima colonna contiene solo la soglia del giocatore.
    Se il file esiste già, aggiunge una nuova riga.
    Se il file non esiste, lo crea con intestazioni.
    
    soglia_giocatore: valore della soglia per questa simulazione
    dati: dizionario con chiavi come intestazioni e valori numerici
    """
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistiche"
        
        # Intestazioni, prima colonna = Soglia Giocatore
        ws.cell(row=1, column=1, value="Soglia Giocatore").font = Font(bold=True)
        for col_idx, intestazione in enumerate(dati.keys(), start=2):
            cella = ws.cell(row=1, column=col_idx, value=intestazione)
            cella.font = Font(bold=True)
            cella.alignment = Alignment(horizontal="center")
    
    # Trova prossima riga vuota
    next_row = ws.max_row + 1

    # Prima colonna: soglia giocatore
    ws.cell(row=next_row, column=1, value=soglia_giocatore)

    # Inserisci i dati nelle colonne successive
    for col_idx, key in enumerate(dati.keys(), start=2):
        ws.cell(row=next_row, column=col_idx, value=dati[key])

    wb.save(file_path)
